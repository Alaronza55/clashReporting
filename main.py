from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import calendar
import time
from datetime import datetime 
import os
import openpyxl
from openpyxl import Workbook
from xlsxwriter.workbook import Workbook
import tkinter as tk
import tkinter.messagebox as msgbox
import pyautogui
import win32com.client as win32
from pywintypes import com_error
import win32com.client
import fitz

#Variables
FILENAME = "XXX"
NEW_NAME = "XXX.xlsx"
current_GMT = time.gmtime()
USERS = os.path.expanduser('~')
FOLDERNAME='Extract'
Downloads=(f'{USERS}\Downloads')
Documents=(f'{USERS}\Documents')
Extract = (f'{Documents}\{FOLDERNAME}')
REPORT = "REPORT.xlsx"

#Get screen size
def extract_screen_width_height():
    screen_width, screen_height = pyautogui.size()
    return screen_width, screen_height

screen_width, screen_height = extract_screen_width_height()
x = (screen_width // 2)-200 
y = (screen_height // 2)-200

def starting_app() :
    if email_entry_Autodesk.get() == "" :
        root.destroy()
        exit()
    elif password_entry_Autodesk.get() =="":
        root.destroy()
        exit()
    elif email_entry_Outlook.get() =="":
        root.destroy()
        exit()
    elif password_entry_Outlook.get() =="":
        root.destroy()
        exit()
    elif indice_entry.get() =="":
        root.destroy()
        exit()
    else:
        get_input()


## GUI to get USER Input from ADESK BIM360
def get_input():
    emailADSK = email_entry_Autodesk.get()
    passwordADSK = password_entry_Autodesk.get()
    emailOutlook = email_entry_Outlook.get()
    passwordOutlook = password_entry_Outlook.get()
    indice = indice_entry.get()
    user_info = {
      "email Autodesk": emailADSK,
      "password Autodesk": passwordADSK,
      "email Outlook": emailOutlook,
      "password Outlook": passwordOutlook,
      "indice" : indice
    }
    root.destroy()
    if root.destroy : True 
    
    def delete_oldies():
        os.chdir(Downloads) 
        for file_name in os.listdir(Downloads):        
            if file_name.startswith("XXX"):  
                os.remove(os.path.join(Downloads, file_name))
                
    #Delete all TEMP files
    print('Deleting all old files from Downloads directory')
    delete_oldies()

    #Webdriver config:
    service_obj = Service("WebDrivers_path\chromedriver.exe")
    driver = webdriver.Chrome(service=service_obj)

    driver.maximize_window()
    driver.get("https://docs.b360.autodesk.com/projects/fdf9236c-9b37-4092-8f87-25f1929658e4/issues")
    time.sleep(40)

    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/form/div/div[1]/div[2]/div/div[1]/div/input').send_keys(emailADSK)
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/form/div/div[1]/div[3]/div/button').click()
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/form/div/div[2]/div[3]/div/div/div[1]/input').send_keys(passwordADSK)
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/form/div/div[2]/div[4]/div/button/span').click()
    time.sleep(40)

    #In BIM360 - ISSUES
    driver.find_element(By.CSS_SELECTOR, '#ProjectIssuesContainer > div > div.ProjectView > div.ProjectGrid > div.ProjectGridHeader > div.ProjectGridHeader__left-buttons > button.Button.Button--icon.FilterIcon').click()
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div[10]/div/div[2]/div[2]/div[2]/div/div[1]/div/div[2]/div/div[2]/div/div/label/div').click()
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div[10]/div/div[2]/div[2]/div[2]/div/div[1]/div/div[2]/div/div[2]/div/div/div/ul/li[8]/div/div/div/span').click()
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div[10]/div/div[2]/div[2]/div[1]/div[2]/div/div/div/button/span[1]').click()
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div[10]/div/div[2]/div[2]/div[1]/div[2]/div/div/div/ul/li[2]/div').click()

    #Download new report
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div[10]/div/div[2]/div[2]/div[3]/div[2]/div/div/div[2]/div/div[1]/div/div/input').send_keys(FILENAME)
    time.sleep(10)
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div[10]/div/div[2]/div[2]/div[3]/div[2]/div/div/div[3]/div/button[2]').click()

    print("Connecting to Outlook and downloading report...")

    #Outlook

    driver.maximize_window()
    driver.get("https://outlook.office365.com/mail/AAMkADNjMDFhMDljLTcxNDItNDU1My04ZWJkLWE3MjY3YzQyMWE4NgAuAAAAAABtZmj7uFKxTJUDw1rK%2B2UaAQClX%2BsphslyRLVqrG8fet%2FsAAAQ6hlfAAA%3D")
    time.sleep(40)

    driver.find_element(By.XPATH, '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div[1]/div[3]/div/div/div/div[2]/div[2]/div/input[1]').send_keys(emailOutlook)
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div[1]/div[3]/div/div/div/div[4]/div/div/div/div/input').click()
    time.sleep(5)

    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/div[2]/div/div/form/div[2]/div[2]/input').send_keys(passwordOutlook)
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[1]/div[2]/div/div/form/div[2]/div[4]/span').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div/form/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[3]/div[2]/div/div/div[1]/input').click()
    time.sleep(20)

    driver.find_element(By.XPATH, '/html/body/div[2]/div/div[2]/div[2]/div[2]/div/div/div/div[1]/div/div/div/div/div[3]/div[16]/div').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '/html/body/div[2]/div/div[2]/div[2]/div[2]/div/div/div/div[3]/div/div[3]/div[1]/div[2]/div/div/div/div/div/div[1]/div/div/div[1]/div[1]/div').click()
    time.sleep(20)

    HTML = str(driver.find_element(By.CSS_SELECTOR, '#ReadingPaneContainerId > div > div > div > div.L72vd > div > div > div.aVla3 > div > div > div > div > div.XbIp4.jmmB7.GNqVo.yxtKT.allowTextSelection > div > div > div > div > div.x_content-wrapper > table > tbody > tr > td > div:nth-child(3) > table > tbody > tr > td > div:nth-child(7) > table > tbody > tr > td > div > table > tbody > tr > td > table > tbody > tr > td > a').get_attribute('outerHTML'))

    print(HTML)

    x = str(HTML.split("\"")[1])
    print(x)

    driver.get(x)

    time.sleep(30)

    def timestamp() :
        current_GMT = time.gmtime()

        time_stamp = calendar.timegm(current_GMT)
        print("Current timestamp:", time_stamp)

        date_time = datetime.fromtimestamp(time_stamp)
        print("The date and time is:", date_time)

        formatted_date_time = date_time.strftime("%Y_%m_%d_%H_%M_%S")

        reportName = f"Report_{formatted_date_time}.xlsx"

        return formatted_date_time

    timestamp()

    def rename_Downloads():
        os.chdir(Downloads) 
        for file in os.listdir(Downloads):        
            if file.startswith("XXX"): 
                old_name = str(file)
                old_name_path = os.path.join(Downloads,old_name)
                new_name_path = os.path.join(Downloads,NEW_NAME)
                os.rename(old_name_path,new_name_path)

    rename_Downloads()

    def wb_Treatement() :
        os.chdir(Downloads) 
        wb = openpyxl.load_workbook("XXX.xlsx") 
        Overview = wb['Overview']
        Issues = wb['Issues']
        last_row = Issues.max_row

        wb.remove(Overview)
        sheet_obj = wb.active 
        Issues.insert_cols(2)

        for i in range(2,last_row+1):
            Issues = wb['Issues']
        Issues.cell(row=i,column=1).hyperlink.target
        # print(Issues.cell(row=i, column=1).hyperlink.target)

        for i in range(2,last_row+1):
            Issues.cell(row=i,column=2).value = Issues.cell(row=i,column=1).hyperlink.target

        print("Saving new reporting source file in Extract directory...")

        os.chdir(Extract)

        wb.save(REPORT)
    
    if os.path.exists(Extract):
        wb_Treatement()
    else :
        os.makedirs(Extract)
        wb_Treatement()

    def deleting_Temp_XXX():

        print("Deleting temp files from Downloads directory")

        os.chdir(Downloads)

        for dirpath, dirnames, filenames in os.walk("."):
            for filename in [f for f in filenames if f.endswith(".xlsx")]:
                print(os.path.join(dirpath, filename))

        for file in os.listdir():
            if file.startswith("XXX"): 
                os.remove('.\XXX.xlsx')

        print("Operation completed. You can now close the program. Thank you come again!")

        msgbox.showinfo("User Info", "File has been saved in : " f'{Extract}')

    deleting_Temp_XXX()

    def open_refresh():
        # Open Excel Application
        excel = win32.gencache.EnsureDispatch('Excel.Application')

        # Open the workbook
        workbook = excel.Workbooks.Open('C:\\Users\\ADavidson\\Documents\\Extract\\IPW1 - TEMPLATE - CLASH DETECTION_V65.xlsx')

        # Refresh all data connections
        workbook.RefreshAll()

        time.sleep(20)

        # Close the workbook and quit Excel Application
        workbook.Close(SaveChanges=True)
        excel.Quit()
    
    open_refresh()

    pdf_Name = f'IPW1-CALSH DETECTION-#{indice}-{timestamp()}.pdf' 

    def rename_header_excel():
        # Open Excel Application
        excel = win32.gencache.EnsureDispatch('Excel.Application')

        # Open the workbook
        workbook = excel.Workbooks.Open('C:\\Users\\ADavidson\\Documents\\Extract\\IPW1 - TEMPLATE - CLASH DETECTION_V65.xlsx')

        excel.Visible = False

        ws = workbook.Worksheets('PVT CHART')

        ws.Cells(1,2).Value = (pdf_Name)

        workbook.Close(SaveChanges=True)
        excel.Quit()
    
    rename_header_excel()

    def export_pdf():
        # Path to original excel file
        WB_PATH = r'C:\Users\ADavidson\Documents\Extract\IPW1 - TEMPLATE - CLASH DETECTION_V65.xlsx'
        # PDF path when saving
        PATH_TO_PDF = r'C:\Users\ADavidson\Documents\Extract\Test.pdf'
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        try:
            print('Start conversion to PDF')
            # Open
            wb = excel.Workbooks.Open(WB_PATH)
            # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
            ws_index_list = [1]
            # Save
            wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
        except com_error as e:
            print('failed.')
        else:
            print('Succeeded.')
        finally:
            wb.Close()
            excel.Quit()

    export_pdf()

    def crop_pdf(input_file, output_file, left, bottom, right, top):
        os.chdir(Extract)
        # Open input PDF file
        with fitz.open(input_file) as pdf:
            # Get the first page of PDF file
            page = pdf[0]
            # Set the crop box for the page
            rect = fitz.Rect(left, bottom, right, top)
            page.set_cropbox(rect)

            # Save the cropped PDF to the output file
            pdf.save(output_file)

    crop_pdf('Test.pdf', pdf_Name, 0, 50, 9300, 5000)

    def remove_test_pdf():
        os.chdir(Extract)
        os.remove('test.pdf')

    remove_test_pdf()



root = tk.Tk()
root.configure(bg='#1d1d1d')
root.geometry(f"+{x}+{y}")
img = tk.PhotoImage(file='.\Pictures\Logo Valens NB.png')
label = tk.Label(root, image=img, bg='#1d1d1d')
label.pack(side='bottom', fill='both', expand='yes')

# Autodesk

root.title("Login: ")

email_label_Autodesk = tk.Label(root, text="Autodesk Account Email:", bg='#1d1d1d',fg='white', font="Arial 12")
email_label_Autodesk.pack()

email_entry_Autodesk = tk.Entry(root, width= 35)
email_entry_Autodesk.pack()

password_label_Autodesk = tk.Label(root, text="Autodesk Account Password:", bg='#1d1d1d',fg='white', font="Arial 12")
password_label_Autodesk.pack()

password_entry_Autodesk = tk.Entry(root, show="*", width= 35)
password_entry_Autodesk.pack()


# Outlook

email_label_Outlook = tk.Label(root, text="Outlook Account Email:", bg='#1d1d1d',fg='white', font="Arial 12")
email_label_Outlook.pack()

email_entry_Outlook = tk.Entry(root, width= 35)
email_entry_Outlook.pack()

password_label_Outlook = tk.Label(root, text="Outlook Account Password:", bg='#1d1d1d',fg='white', font="Arial 12")
password_label_Outlook.pack()

password_entry_Outlook = tk.Entry(root, show="*", width= 35)
password_entry_Outlook.pack()

# Indice

indice_Label = tk.Label(root, text="Indice du rapport:", bg='#1d1d1d',fg='white', font="Arial 12")
indice_Label.pack()

indice_entry = tk.Entry(root, width= 35)
indice_entry.pack()

submit_button = tk.Button(root, text="Submit", command=starting_app, bg='#1d1d1d',fg='white', font="Arial 12",border='0' , padx=10, pady=20)
submit_button.pack()

root.mainloop()

time.sleep(30)